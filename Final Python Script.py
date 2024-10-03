import pyodbc
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, Border, Side

server = 'ABHI-PC\SQLEXPRESS'
database = 'AzureDE_Project'
username = 'AbhiDEProjectUserOne'
password = 'abcd9876'

conn = pyodbc.connect(f'DRIVER={{ODBC Driver 17 for SQL Server}};'
                      f'SERVER={server};DATABASE={database};'
                      f'UID={username};PWD={password}')

# Function to get schema of a table
def get_table_schema(table_name):
    schema_query = f"""
    SELECT COLUMN_NAME, 
    CASE 
        WHEN CHARACTER_MAXIMUM_LENGTH IS NOT NULL THEN 
            DATA_TYPE + '(' + CAST(CHARACTER_MAXIMUM_LENGTH AS VARCHAR) + ')'
        ELSE DATA_TYPE
    END AS DATA_TYPE
    ,IS_NULLABLE
    FROM INFORMATION_SCHEMA.COLUMNS
    WHERE TABLE_NAME = '{table_name}'
    """
    return pd.read_sql(schema_query, conn)

# Function to get data from a table
def get_table_data(table_name):
    data_query = f"SELECT * FROM {table_name}"
    return pd.read_sql(data_query, conn)


tables_query = "SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_TYPE = 'BASE TABLE'"
tables_df = pd.read_sql(tables_query, conn)
table_names = tables_df['TABLE_NAME'].tolist()

## select the number of tables needed in the output
table_names = table_names[:2]

tables = {}

# Loop through each table and store its schema and data in the dictionary
for table in table_names:
    tables[table] = {
        'schema': get_table_schema(table),
        #'data': get_table_data(table)
    }



# Initialize ExcelWriter to store data in separate sheets
with pd.ExcelWriter('first_2_tables_schema_and_data.xlsx', engine='openpyxl') as writer:
    for table in table_names:
        # Get schema and data for each table
        table_schema = get_table_schema(table)
        #table_data = get_table_data(table)
        
        # Write schema and data to separate sheets in Excel
        table_schema.to_excel(writer, sheet_name=f'{table}_schema', index=False)
        #table_data.to_excel(writer, sheet_name=f'{table}_data', index=False)


wb = load_workbook('first_2_tables_schema_and_data.xlsx')

# Define border style
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


# Function to apply borders and formatting to a sheet
def format_sheet(sheet):
    # Apply formatting to header row
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        # Optionally, set column widths based on content
        sheet.column_dimensions[cell.column_letter].width = 20
    
    # Apply borders to all data cells
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=sheet.max_column):
        for cell in row:
            cell.border = thin_border
            # Optional: Set alignment for data cells
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)


# Loop through the first 3 tables for formatting
for table in table_names:
    # Access the schema and data sheets
    schema_sheet = wb[f'{table}_schema']
    # Example: Add a custom column "test_column" to schema sheet
    # description_col = schema_sheet.max_column + 1
   
    schema_sheet.insert_cols(2)
    schema_sheet.cell(row=1, column=2).value = "Description"
    
    schema_sheet.insert_cols(4)
    schema_sheet.cell(row=1, column=4).value = "UK/PK/FK"
    
    schema_sheet.insert_cols(5)
    schema_sheet.cell(row=1, column=5).value = "is_Required - Sliver"

    schema_sheet.insert_cols(6)
    schema_sheet.cell(row=1, column=6).value = "Transformation required"
    
    schema_sheet.insert_cols(8)
    schema_sheet.cell(row=1, column=8).value = "Example"
    
    
    
    # schema_sheet.cell(row=1, column=description_col).value = "test_column"
    # schema_sheet.column_dimensions[schema_sheet.cell(row=1, column=description_col).column_letter].width = 30
    
    # Apply formatting to schema sheet
    format_sheet(schema_sheet)

# Save the formatted workbook
formatted_filename = 'first_2_tables_schema_formatted_with_borders.xlsx'
wb.save(formatted_filename)

